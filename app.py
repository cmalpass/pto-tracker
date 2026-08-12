"""PTO Tracker - Paid Time Off Calculator & Forecaster"""
import os
import json
import sqlite3
import math
import calendar
import hmac
import secrets
import logging
from datetime import datetime, timedelta, date
from flask import Flask, jsonify, request, render_template, g
import holidays

app = Flask(__name__)
CSRF_COOKIE_NAME = 'pto_csrf_token'
CSRF_HEADER_NAME = 'X-CSRF-Token'
API_AUTH_SCHEME = 'Bearer'
DATABASE = os.environ.get(
    'PTO_DB_PATH',
    os.path.join(os.path.dirname(__file__), 'instance', 'pto_tracker.db')
)
logger = logging.getLogger(__name__)

NUMERIC_CONFIG_KEYS = {
    'pto_accrual_per_pay_period',
    'pto_hours_per_day',
    'pay_periods_per_year',
    'pto_carryover_limit',
    'pto_cashout_rate',
    'pto_grace_period_days',
}
BOOLEAN_CONFIG_KEYS = {
    'pto_holidays_require_pto',
    'pto_uses_rollover',
    'pto_lose_above_limit',
}
VALID_CONFIG_KEYS = {
    'pto_accrual_per_pay_period',
    'pto_accrual_type',
    'pto_hours_per_day',
    'pto_holidays_require_pto',
    'pay_periods_per_year',
    'accrual_start_date',
    'accrual_method',
    'pto_carryover_limit',
    'pto_uses_rollover',
    'pto_cashout_rate',
    'pto_lose_above_limit',
    'pto_start_year',
    'pto_vesting_schedule',
    'pto_grace_period_days',
}


def default_config():
    return {
        'pto_accrual_per_pay_period': '1.0',
        'pto_accrual_type': 'days',
        'pto_hours_per_day': '8',
        'pto_holidays_require_pto': 'true',
        'pay_periods_per_year': '26',
        'accrual_start_date': f'{datetime.now().year}-01-01',
        'accrual_method': 'pro-rata',
        'pto_carryover_limit': '40',
        'pto_uses_rollover': 'true',
        'pto_cashout_rate': '0',
        'pto_lose_above_limit': 'true',
        'pto_start_year': str(datetime.now().year),
        'pto_vesting_schedule': 'immediate',
        'pto_grace_period_days': '0',
    }


def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
        g.db.execute('PRAGMA journal_mode=WAL')
    return g.db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop('db', None)
    if db is not None:
        db.close()


def init_db():
    database_dir = os.path.dirname(DATABASE)
    if database_dir:
        os.makedirs(database_dir, exist_ok=True)
    conn = sqlite3.connect(DATABASE)
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS config (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS vacations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            days REAL NOT NULL,
            hours REAL NOT NULL DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS notes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            text TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
    ''')
    defaults = default_config()
    for key, value in defaults.items():
        conn.execute('INSERT OR IGNORE INTO config (key, value) VALUES (?, ?)',
                     (key, value))
    conn.commit()
    conn.close()


def get_config():
    db = get_db()
    rows = db.execute('SELECT key, value FROM config').fetchall()
    config = {}
    for row in rows:
        key = row['key']
        value = row['value']
        if key in NUMERIC_CONFIG_KEYS:
            try:
                config[key] = float(value)
            except (TypeError, ValueError):
                fallback = float(default_config()[key])
                logger.warning('Invalid numeric config %s=%r; using %s', key, value, fallback)
                config[key] = fallback
        elif key in BOOLEAN_CONFIG_KEYS:
            config[key] = value.lower() == 'true'
        else:
            config[key] = value
    return config


def has_valid_basic_auth():
    auth = request.authorization
    expected_username = os.getenv('PTO_AUTH_USERNAME', '')
    expected_password = os.getenv('PTO_AUTH_PASSWORD', '')
    return bool(
        auth
        and expected_username
        and expected_password
        and auth.username
        and auth.password
        and hmac.compare_digest(auth.username, expected_username)
        and hmac.compare_digest(auth.password, expected_password)
    )


@app.before_request
def require_auth_for_writes():
    if request.method not in {'POST', 'PUT', 'DELETE'}:
        return None
    if os.getenv('PTO_REQUIRE_AUTH', 'false').strip().lower() not in {'1', 'true', 'yes', 'on'}:
        return None

    if not has_valid_basic_auth():
        response = jsonify({'error': 'Authentication required'})
        response.status_code = 401
        response.headers['WWW-Authenticate'] = 'Basic realm="PTO Tracker"'
        return response
    return None


def has_valid_api_auth():
    authorization = request.headers.get('Authorization', '')
    scheme, _, credential = authorization.partition(' ')
    expected_api_key = os.getenv('PTO_API_KEY', '')
    if (
        scheme.lower() != API_AUTH_SCHEME.lower()
        or not credential
        or not expected_api_key
    ):
        return False
    return hmac.compare_digest(credential, expected_api_key)


@app.before_request
def require_csrf_for_browser_writes():
    if request.method not in {'POST', 'PUT', 'DELETE', 'PATCH'}:
        return None

    cookie_token = request.cookies.get(CSRF_COOKIE_NAME)
    if cookie_token is None:
        # Cookie-less writes must identify themselves as API clients explicitly.
        if has_valid_api_auth() or (
            os.getenv('PTO_REQUIRE_AUTH', 'false').strip().lower()
            in {'1', 'true', 'yes', 'on'}
            and has_valid_basic_auth()
        ):
            return None
        return jsonify({'error': 'CSRF validation failed'}), 403

    header_token = request.headers.get(CSRF_HEADER_NAME, '')
    if not header_token or not hmac.compare_digest(header_token, cookie_token):
        return jsonify({'error': 'CSRF validation failed'}), 403
    return None


@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src https://fonts.gstatic.com; "
        "connect-src 'self'; "
        "base-uri 'self'; frame-ancestors 'none'"
    )
    return response


def get_us_holidays(year):
    return holidays.US(years=year, observed=True)


def is_business_day(date):
    return date.weekday() < 5


def get_vacation_days(start_date, end_date, config):
    start = datetime.strptime(start_date, '%Y-%m-%d').date()
    end = datetime.strptime(end_date, '%Y-%m-%d').date()
    holidays_require_pto = config.get('pto_holidays_require_pto', True)
    holidays_set = set()
    if not holidays_require_pto:
        for year in range(start.year, end.year + 1):
            holidays_set.update(get_us_holidays(year))
    days = 0
    current = start
    while current <= end:
        if is_business_day(current) and (holidays_require_pto or current not in holidays_set):
            days += 1
        current += timedelta(days=1)
    return days


def parse_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in ('1', 'true', 'yes', 'on')
    return bool(value)


def normalize_quarter_hours(value):
    try:
        hours = float(value or 0)
    except (TypeError, ValueError):
        return None
    if hours < 0:
        return None
    return round(hours * 4) / 4


def calculate_accrual_to_date(target_date, config):
    accrual_start = datetime.strptime(config['accrual_start_date'], '%Y-%m-%d').date()
    target = datetime.strptime(target_date, '%Y-%m-%d').date()
    if target < accrual_start:
        return 0.0
    pay_period_days = 365.25 / config['pay_periods_per_year']
    days_elapsed = (target - accrual_start).days
    if config['accrual_method'] == 'pro-rata':
        holidays_set = set()
        for y in range(accrual_start.year, target.year + 1):
            holidays_set.update(get_us_holidays(y))
        business_days_worked = _count_business_days(accrual_start, target, holidays_set)
        accrual_per_day = config['pto_accrual_per_pay_period'] / (pay_period_days * 5 / 7)
        accrued = business_days_worked * accrual_per_day
    else:
        pay_periods_elapsed = days_elapsed / pay_period_days
        accrued = pay_periods_elapsed * config['pto_accrual_per_pay_period']
    return accrued


def _count_business_days(start, end, holidays_set=None):
    """Count weekdays in an inclusive date range, excluding supplied holidays."""
    if start > end:
        return 0
    total_days = (end - start).days + 1
    full_weeks, remainder = divmod(total_days, 7)
    business_days = full_weeks * 5
    business_days += sum(
        (start.weekday() + offset) % 7 < 5
        for offset in range(remainder)
    )
    if holidays_set:
        business_days -= sum(
            start <= holiday <= end and holiday.weekday() < 5
            for holiday in holidays_set
        )
    return business_days


def calculate_vacation_usage_in_range(range_start, range_end, config):
    """Return (days, hours) of vacation used within [range_start, range_end] (date objects).

    Vacations that overlap the range are clipped to it. Year-spanning vacations are
    split so each year only receives the days that fall within it.
    """
    db = get_db()
    rows = db.execute(
        'SELECT * FROM vacations WHERE start_date <= ? AND end_date >= ? ORDER BY start_date',
        (range_end.strftime('%Y-%m-%d'), range_start.strftime('%Y-%m-%d'))
    ).fetchall()
    total_days = 0.0
    total_hours = 0.0
    for row in rows:
        vac_start = datetime.strptime(row['start_date'], '%Y-%m-%d').date()
        vac_end = datetime.strptime(row['end_date'], '%Y-%m-%d').date()
        eff_start = max(vac_start, range_start)
        eff_end = min(vac_end, range_end)
        if eff_start > eff_end:
            continue
        if eff_start == vac_start and eff_end == vac_end:
            # Entirely within range — use stored (authoritative) values
            total_days += row['days']
            total_hours += row['hours']
        else:
            # Partially within range: apportion stored row['days'] by overlap ratio
            # so split ranges never consume more than the saved entry total.
            total_business_days = get_vacation_days(
                vac_start.strftime('%Y-%m-%d'),
                vac_end.strftime('%Y-%m-%d'),
                config
            )
            overlap_business_days = get_vacation_days(
                eff_start.strftime('%Y-%m-%d'),
                eff_end.strftime('%Y-%m-%d'),
                config
            )

            if total_business_days > 0:
                total_days += row['days'] * (overlap_business_days / total_business_days)
            elif vac_start >= range_start:
                # Degenerate edge case: no business days in range. Attribute manual
                # override to the range containing the start date.
                total_days += row['days']

            # Partial-hour PTO is always a single day; only attribute hours when
            # the vacation's start date falls inside this range.
            if vac_start >= range_start:
                total_hours += row['hours']
    return total_days, total_hours


def calculate_balance_on_date(target_date, config):
    """Compute PTO balance on target_date with proper year-end rollover.

    For the first accrual year: balance = accrued − used (no cap during year).
    For subsequent years:
      1. Compute each prior year's Dec-31 balance iteratively.
      2. Apply rollover rules:
         - pto_uses_rollover=False  → carry = 0 (use-it-or-lose-it)
         - pto_lose_above_limit=True → carry = min(carry, carryover_limit)
         - otherwise               → carry = full Dec-31 balance
      3. balance = carry + this_year_accruals − this_year_usage
    The cap (lose_above_limit) is enforced ONLY at year-end, not continuously.
    """
    target = datetime.strptime(target_date, '%Y-%m-%d').date()
    accrual_start = datetime.strptime(config['accrual_start_date'], '%Y-%m-%d').date()

    hours_per_day = config.get('pto_hours_per_day', 8.0) or 8.0
    is_hours = config.get('pto_accrual_type') == 'hours'
    carryover_limit_days = config['pto_carryover_limit']
    uses_rollover = config.get('pto_uses_rollover', True)
    lose_above_limit = config.get('pto_lose_above_limit', False)
    effective_limit = (carryover_limit_days * hours_per_day) if is_hours else carryover_limit_days

    if target < accrual_start:
        return {
            'accrued': 0.0, 'used': 0.0, 'used_days': 0.0, 'used_hours': 0.0,
            'balance': 0.0, 'limit': round(effective_limit, 2), 'carry': 0.0
        }

    # Walk each accrual year once, carrying forward the prior year's ending balance.
    carry_balance = 0.0
    for year in range(accrual_start.year, target.year + 1):
        year_window_start = max(accrual_start, date(year, 1, 1))
        year_window_end = min(target, date(year, 12, 31))
        year_accrual = (
            calculate_accrual_to_date(year_window_end.strftime('%Y-%m-%d'), config)
            - calculate_accrual_to_date(
                (year_window_start - timedelta(days=1)).strftime('%Y-%m-%d'), config
            )
        )
        used_days, used_hours = calculate_vacation_usage_in_range(
            year_window_start, year_window_end, config
        )
        if is_hours:
            used_amount = (used_days * hours_per_day) + used_hours
        else:
            used_amount = used_days + (used_hours / hours_per_day)

        balance = max(0.0, carry_balance + year_accrual - used_amount)
        if year == target.year:
            return {
                'accrued': round(carry_balance + year_accrual, 2),
                'used': round(used_amount, 2),
                'used_days': round(used_days, 2),
                'used_hours': round(used_hours, 2),
                'balance': round(balance, 2),
                'limit': round(effective_limit, 2),
                'carry': round(carry_balance, 2)
            }

        carry_balance = balance
        if not uses_rollover:
            carry_balance = 0.0
        elif lose_above_limit and carry_balance > effective_limit:
            carry_balance = effective_limit


def generate_yearly_forecast(year, config):
    forecast = []
    for month in range(1, 13):
        if month == 12:
            end_date = datetime(year, 12, 31)
        else:
            end_date = datetime(year, month + 1, 1) - timedelta(days=1)
        balance = calculate_balance_on_date(end_date.strftime('%Y-%m-%d'), config)
        balance['month'] = end_date.strftime('%Y-%m')
        balance['month_name'] = end_date.strftime('%B')
        forecast.append(balance)
    return forecast


def generate_calendar_events(year, config):
    events = []
    us_holidays = get_us_holidays(year)
    for holiday_date, name in sorted(us_holidays.items()):
        events.append({
            'date': holiday_date.strftime('%Y-%m-%d'),
            'type': 'holiday',
            'name': name,
            'color': '#e74c3c'
        })
    db = get_db()
    rows = db.execute(
        'SELECT * FROM vacations WHERE start_date <= ? AND end_date >= ?',
        (f'{year}-12-31', f'{year}-01-01')
    ).fetchall()
    for row in rows:
        start = datetime.strptime(row['start_date'], '%Y-%m-%d').date()
        end = datetime.strptime(row['end_date'], '%Y-%m-%d').date()
        start = max(start, date(year, 1, 1))
        end = min(end, date(year, 12, 31))
        for day in _daterange(start, end):
            events.append({
                'date': day.strftime('%Y-%m-%d'),
                'type': 'vacation',
                'name': row['name'],
                'color': '#3498db',
                'vacation_id': row['id']
            })
    events.sort(key=lambda x: x['date'])
    return events


def validate_vacation_name(value):
    name = ''.join(
        character for character in str(value or 'Vacation').strip()
        if character.isprintable()
    )[:100]
    return name or 'Vacation'


def _booking_amount(days, hours, config):
    hours_per_day = config.get('pto_hours_per_day', 8.0) or 8.0
    if config.get('pto_accrual_type') == 'hours':
        return (days * hours_per_day) + hours
    return days + (hours / hours_per_day)


def _existing_booking_amount_through(existing, target, config):
    existing_start = datetime.strptime(existing['start_date'], '%Y-%m-%d').date()
    existing_end = datetime.strptime(existing['end_date'], '%Y-%m-%d').date()
    accrual_start = datetime.strptime(config['accrual_start_date'], '%Y-%m-%d').date()
    year_start = accrual_start if target.year <= accrual_start.year else date(target.year, 1, 1)
    overlap_start = max(existing_start, year_start)
    overlap_end = min(existing_end, target)
    if overlap_start > overlap_end:
        return 0.0

    total_business_days = get_vacation_days(
        existing['start_date'], existing['end_date'], config
    )
    overlap_business_days = get_vacation_days(
        overlap_start.strftime('%Y-%m-%d'),
        overlap_end.strftime('%Y-%m-%d'),
        config
    )
    if total_business_days > 0:
        days = existing['days'] * (overlap_business_days / total_business_days)
    elif existing_start >= year_start:
        days = existing['days']
    else:
        days = 0.0
    hours = existing['hours'] if existing_start >= year_start else 0.0
    return _booking_amount(days, hours, config)


def _validate_booking_balance(end_date, days, hours, config, existing=None):
    projected = calculate_balance_on_date(end_date, config)
    available = projected['accrued'] - projected['used']
    if existing:
        target = datetime.strptime(end_date, '%Y-%m-%d').date()
        available += _existing_booking_amount_through(existing, target, config)
    requested = _booking_amount(days, hours, config)
    if available - requested < -1e-9:
        unit = 'hours' if config.get('pto_accrual_type') == 'hours' else 'days'
        return (
            f'Vacation would create a negative balance: '
            f'{available - requested:.2f} {unit} projected after booking'
        )
    return None


def _daterange(start, end):
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def _build_reserved_dates(year, vacations):
    reserved = set()
    year_start = date(year, 1, 1)
    year_end = date(year, 12, 31)
    for row in vacations:
        start = datetime.strptime(row['start_date'], '%Y-%m-%d').date()
        end = datetime.strptime(row['end_date'], '%Y-%m-%d').date()
        overlap_start = max(start, year_start)
        overlap_end = min(end, year_end)
        if overlap_start > overlap_end:
            continue
        for day in _daterange(overlap_start, overlap_end):
            reserved.add(day)
    return reserved


def _valid_pto_day(day, holidays_set, reserved_dates, earliest_date, year, holidays_require_pto):
    return (
        day.year == year
        and day >= earliest_date
        and is_business_day(day)
        and (holidays_require_pto or day not in holidays_set)
        and day not in reserved_dates
    )


def _continuous_days_off_count(pto_dates, holidays_set, min_date=None, max_date=None):
    if not pto_dates:
        return 0

    def is_day_off(check_day):
        return (check_day.weekday() >= 5) or (check_day in holidays_set) or (check_day in pto_dates)

    start = min(pto_dates)
    end = max(pto_dates)
    min_date = min_date or date(start.year, 1, 1)
    max_date = max_date or date(end.year, 12, 31)

    cursor = start - timedelta(days=1)
    while cursor >= min_date and is_day_off(cursor):
        start = cursor
        cursor -= timedelta(days=1)

    cursor = end + timedelta(days=1)
    while cursor <= max_date and is_day_off(cursor):
        end = cursor
        cursor += timedelta(days=1)

    return (end - start).days + 1


def _make_suggestion(start_day, end_day, title, reason, category, holidays_set, holidays_require_pto, holiday_date=None):
    all_dates = list(_daterange(start_day, end_day))
    pto_dates = [
        day.strftime('%Y-%m-%d')
        for day in all_dates
        if day.weekday() < 5 and (holidays_require_pto or day not in holidays_set)
    ]
    pto_days = len(pto_dates)
    days_off = _continuous_days_off_count(
        set(all_dates),
        holidays_set,
        min_date=date(start_day.year, 1, 1),
        max_date=date(start_day.year, 12, 31)
    )
    return {
        'name': title,
        'start_date': start_day.strftime('%Y-%m-%d'),
        'end_date': end_day.strftime('%Y-%m-%d'),
        'holiday_date': holiday_date.strftime('%Y-%m-%d') if holiday_date else None,
        'pto_dates': pto_dates,
        'pto_days': pto_days,
        'total_days_off': days_off,
        'impact_score': round(days_off / pto_days, 2) if pto_days else 0,
        'category': category,
        'reason': reason,
        'tags': [category.replace('-', ' ')]
    }


def generate_vacation_suggestions(year, config):
    db = get_db()
    vacations = db.execute('SELECT * FROM vacations ORDER BY start_date').fetchall()

    today = datetime.now().date()
    year_start = date(year, 1, 1)
    year_end = date(year, 12, 31)
    earliest = max(today, year_start)

    is_hours = config.get('pto_accrual_type') == 'hours'
    hours_per_day = config.get('pto_hours_per_day', 8.0) or 8.0
    holidays_require_pto = config.get('pto_holidays_require_pto', True)

    dec31_balance = calculate_balance_on_date(year_end.strftime('%Y-%m-%d'), config)
    # The balance helper includes all bookings through year-end, so suggestions
    # only budget the amount that remains after existing vacations.
    remaining_amount = max(
        0.0, dec31_balance['accrued'] - dec31_balance['used']
    )
    remaining_days = (remaining_amount / hours_per_day) if is_hours else remaining_amount

    carry_limit_days = config.get('pto_carryover_limit', 0) or 0
    carry_limit_amount = (carry_limit_days * hours_per_day) if is_hours else carry_limit_days
    if not config.get('pto_uses_rollover', True):
        forfeit_amount = remaining_amount
    elif config.get('pto_lose_above_limit', False):
        forfeit_amount = max(0.0, remaining_amount - carry_limit_amount)
    else:
        forfeit_amount = 0.0

    forfeit_days = (forfeit_amount / hours_per_day) if is_hours else forfeit_amount
    target_days = forfeit_days if forfeit_days > 0 else min(remaining_days, 10.0)
    budget_days = max(0, int(math.floor(target_days + 1e-9)))

    reserved_dates = _build_reserved_dates(year, vacations)
    holidays_map = get_us_holidays(year)
    holidays_set = set(holidays_map.keys())

    candidates = []
    seen_ranges = set()

    def add_candidate(start_day, end_day, title, reason, category, holiday_date=None):
        if start_day > end_day:
            return
        if start_day < earliest:
            return
        for d in _daterange(start_day, end_day):
            # Vacation ranges may include a holiday date, but all non-holiday weekdays
            # must be valid PTO days and not already reserved.
            if d == holiday_date:
                continue
            if not _valid_pto_day(d, holidays_set, reserved_dates, earliest, year, holidays_require_pto):
                return

        if holiday_date and holiday_date.year == year and holiday_date >= earliest and holiday_date not in reserved_dates:
            start_day = min(start_day, holiday_date)
            end_day = max(end_day, holiday_date)

        key = (start_day, end_day)
        if key in seen_ranges:
            return
        seen_ranges.add(key)
        suggestion = _make_suggestion(
            start_day,
            end_day,
            title,
            reason,
            category,
            holidays_set,
            holidays_require_pto,
            holiday_date=holiday_date
        )
        if suggestion['pto_days'] <= 0:
            return
        candidates.append(suggestion)

    for holiday_day, holiday_name in sorted(holidays_map.items()):
        weekday = holiday_day.weekday()
        if weekday == 0:
            add_candidate(
                holiday_day - timedelta(days=3),
                holiday_day - timedelta(days=3),
                f'Extend {holiday_name}',
                f'Take Friday off to turn {holiday_name} into a 4-day break.',
                'holiday-bridge',
                holiday_date=holiday_day
            )
        elif weekday == 1:
            add_candidate(
                holiday_day - timedelta(days=1),
                holiday_day - timedelta(days=1),
                f'Bridge into {holiday_name}',
                f'Take Monday off before {holiday_name} for a longer break.',
                'holiday-bridge',
                holiday_date=holiday_day
            )
        elif weekday == 2:
            add_candidate(
                holiday_day + timedelta(days=1),
                holiday_day + timedelta(days=2),
                f'Long break after {holiday_name}',
                f'Take Thu/Fri after {holiday_name} for a 5-day stretch.',
                'holiday-bridge',
                holiday_date=holiday_day
            )
        elif weekday == 3:
            add_candidate(
                holiday_day + timedelta(days=1),
                holiday_day + timedelta(days=1),
                f'Extend {holiday_name}',
                f'Take Friday off after {holiday_name} for a 4-day weekend.',
                'holiday-bridge',
                holiday_date=holiday_day
            )
        elif weekday == 4:
            add_candidate(
                holiday_day + timedelta(days=3),
                holiday_day + timedelta(days=3),
                f'Extend {holiday_name}',
                f'Take Monday off after {holiday_name} for extra recovery time.',
                'holiday-bridge',
                holiday_date=holiday_day
            )

    current = earliest
    while current <= year_end:
        if current.weekday() in (0, 4):
            add_candidate(
                current,
                current,
                'Create a Long Weekend',
                'Use a single PTO day next to the weekend for a 3-day break.',
                'high-impact'
            )
        current += timedelta(days=1)

    candidates.sort(key=lambda c: (c['impact_score'], c['total_days_off']), reverse=True)

    selected = []
    selected_dates = set(reserved_dates)
    used_days = 0
    hard_cap_days = max(0, int(min(max(remaining_days, 0), 15)))
    max_days_to_use = budget_days if budget_days > 0 else hard_cap_days

    for candidate in candidates:
        if len(selected) >= 12:
            break
        candidate_start = datetime.strptime(candidate['start_date'], '%Y-%m-%d').date()
        candidate_end = datetime.strptime(candidate['end_date'], '%Y-%m-%d').date()
        candidate_dates = set(_daterange(candidate_start, candidate_end))
        if candidate_dates & selected_dates:
            continue
        if max_days_to_use > 0 and (used_days + candidate['pto_days']) > max_days_to_use:
            continue
        selected.append(candidate)
        selected_dates.update(candidate_dates)
        used_days += candidate['pto_days']

    suggested_hours = used_days * hours_per_day
    if is_hours:
        planned_suggestion_amount = round(suggested_hours, 2)
    else:
        planned_suggestion_amount = round(float(used_days), 2)

    summary_message = 'Suggestions are optimized for holiday alignment and high impact time off.'
    if forfeit_days > 0:
        summary_message = 'You are on track to forfeit PTO unless you schedule additional time off.'
    elif remaining_days > 0:
        summary_message = 'You still have PTO available; these options maximize time off per PTO day.'

    return {
        'year': year,
        'unit': 'hours' if is_hours else 'days',
        'hours_per_day': round(hours_per_day, 2),
        'remaining_balance': round(remaining_amount, 2),
        'remaining_balance_days_equivalent': round(remaining_days, 2),
        'forfeit_risk': round(forfeit_amount, 2),
        'forfeit_risk_days_equivalent': round(forfeit_days, 2),
        'target_to_plan_days': round(target_days, 2),
        'suggested_pto_amount': planned_suggestion_amount,
        'suggested_pto_days': used_days,
        'summary': {
            'message': summary_message,
            'recommendation': 'Add one or more suggestions to your plan to avoid unused PTO at year end.'
        },
        'suggestions': selected
    }


@app.route('/')
def index():
    response = render_template('index.html')
    response = app.make_response(response)
    if request.cookies.get(CSRF_COOKIE_NAME) is None:
        response.set_cookie(
            CSRF_COOKIE_NAME,
            secrets.token_urlsafe(32),
            httponly=False,
            secure=request.is_secure,
            samesite='Strict',
            path='/',
        )
    return response


@app.route('/api/config', methods=['GET'])
def api_get_config():
    return jsonify(get_config())


@app.route('/api/config', methods=['PUT'])
def api_update_config():
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({'error': 'Request body must be a JSON object'}), 400

    validated = {}
    for key, value in data.items():
        if key not in VALID_CONFIG_KEYS:
            return jsonify({'error': f'Invalid config key: {key}'}), 400
        if key in NUMERIC_CONFIG_KEYS:
            if isinstance(value, bool):
                return jsonify({'error': f'{key} must be numeric'}), 400
            try:
                numeric_value = float(value)
            except (TypeError, ValueError):
                return jsonify({'error': f'{key} must be numeric'}), 400
            if not math.isfinite(numeric_value):
                return jsonify({'error': f'{key} must be numeric'}), 400
            validated[key] = str(value)
        elif key in BOOLEAN_CONFIG_KEYS:
            if isinstance(value, bool):
                validated[key] = str(value).lower()
            elif isinstance(value, str) and value.strip().lower() in {'true', 'false'}:
                validated[key] = value.strip().lower()
            else:
                return jsonify({'error': f'{key} must be boolean'}), 400
        else:
            validated[key] = str(value)

    db = get_db()
    for key, value in validated.items():
        db.execute('INSERT OR REPLACE INTO config (key, value) VALUES (?, ?)',
                   (key, value))
    db.commit()
    return jsonify({'status': 'ok', 'config': get_config()})


@app.route('/api/balance/<date>', methods=['GET'])
def api_get_balance(date):
    config = get_config()
    balance = calculate_balance_on_date(date, config)
    return jsonify({'date': date, **balance})


@app.route('/api/balance', methods=['GET'])
def api_get_balance_range():
    year = request.args.get('year', str(datetime.now().year), type=int)
    config = get_config()
    forecast = generate_yearly_forecast(year, config)
    return jsonify({'year': year, 'forecast': forecast})


@app.route('/api/forecast/<date>', methods=['GET'])
def api_get_forecast(date):
    config = get_config()
    balance = calculate_balance_on_date(date, config)
    return jsonify({'date': date, **balance})


@app.route('/api/vacations', methods=['GET'])
def api_get_vacations():
    db = get_db()
    rows = db.execute('SELECT * FROM vacations ORDER BY start_date').fetchall()
    return jsonify([dict(row) for row in rows])


@app.route('/api/vacations', methods=['POST'])
def api_add_vacation():
    data = request.get_json() or {}
    db = get_db()
    name = validate_vacation_name(data.get('name', 'Vacation'))
    start_date = data.get('start_date')
    end_date = data.get('end_date')
    if not start_date or not end_date:
        return jsonify({'error': 'start_date and end_date are required'}), 400
    if start_date > end_date:
        return jsonify({'error': 'start_date cannot be after end_date'}), 400
    config = get_config()
    try:
        datetime.strptime(start_date, '%Y-%m-%d')
        datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'Dates must use yyyy-mm-dd format'}), 400
    auto_days = parse_bool(data.get('auto_days', True), default=True)
    if auto_days:
        days = get_vacation_days(start_date, end_date, config)
    else:
        try:
            days = float(data.get('days', 0) or 0)
        except (TypeError, ValueError):
            return jsonify({'error': 'days must be numeric'}), 400
        if days < 0:
            return jsonify({'error': 'days cannot be negative'}), 400
    hours = normalize_quarter_hours(data.get('hours', 0))
    if hours is None:
        return jsonify({'error': 'hours must be non-negative and in numeric format'}), 400
    balance_error = _validate_booking_balance(
        end_date, days, hours, config
    )
    if balance_error:
        return jsonify({'error': balance_error}), 400
    cursor = db.execute(
        'INSERT INTO vacations (name, start_date, end_date, days, hours) VALUES (?, ?, ?, ?, ?)',
        (name, start_date, end_date, days, hours)
    )
    db.commit()
    row = db.execute('SELECT * FROM vacations WHERE id = ?', (cursor.lastrowid,)).fetchone()
    return jsonify(dict(row)), 201


@app.route('/api/vacations/calculate-days', methods=['GET'])
def api_calculate_vacation_days():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    if not start_date or not end_date:
        return jsonify({'error': 'start_date and end_date are required'}), 400
    if start_date > end_date:
        return jsonify({'error': 'start_date cannot be after end_date'}), 400
    try:
        days = get_vacation_days(start_date, end_date, get_config())
    except ValueError:
        return jsonify({'error': 'Dates must use yyyy-mm-dd format'}), 400
    return jsonify({'start_date': start_date, 'end_date': end_date, 'days': days})


@app.route('/api/vacations/suggestions', methods=['GET'])
def api_get_vacation_suggestions():
    year = request.args.get('year', datetime.now().year, type=int)
    if year < 2000 or year > 2100:
        return jsonify({'error': 'year must be between 2000 and 2100'}), 400
    config = get_config()
    payload = generate_vacation_suggestions(year, config)
    return jsonify(payload)


@app.route('/api/vacations/<int:vacation_id>', methods=['PUT'])
def api_update_vacation(vacation_id):
    data = request.get_json() or {}
    db = get_db()
    existing = db.execute('SELECT * FROM vacations WHERE id = ?', (vacation_id,)).fetchone()
    if not existing:
        return jsonify({'error': 'Vacation not found'}), 404

    name = validate_vacation_name(data.get('name', existing['name']))
    start_date = data.get('start_date', existing['start_date'])
    end_date = data.get('end_date', existing['end_date'])

    if not start_date or not end_date:
        return jsonify({'error': 'start_date and end_date are required'}), 400
    if start_date > end_date:
        return jsonify({'error': 'start_date cannot be after end_date'}), 400
    try:
        datetime.strptime(start_date, '%Y-%m-%d')
        datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'Dates must use yyyy-mm-dd format'}), 400

    config = get_config()
    auto_days = parse_bool(data.get('auto_days', True), default=True)
    if auto_days:
        days = get_vacation_days(start_date, end_date, config)
    else:
        try:
            days = float(data.get('days', existing['days']) or 0)
        except (TypeError, ValueError):
            return jsonify({'error': 'days must be numeric'}), 400
        if days < 0:
            return jsonify({'error': 'days cannot be negative'}), 400
    hours = normalize_quarter_hours(data.get('hours', existing['hours']))
    if hours is None:
        return jsonify({'error': 'hours must be non-negative and in numeric format'}), 400
    balance_error = _validate_booking_balance(
        end_date, days, hours, config, existing=existing
    )
    if balance_error:
        return jsonify({'error': balance_error}), 400

    db.execute(
        'UPDATE vacations SET name = ?, start_date = ?, end_date = ?, days = ?, hours = ? WHERE id = ?',
        (name, start_date, end_date, days, hours, vacation_id)
    )
    db.commit()
    row = db.execute('SELECT * FROM vacations WHERE id = ?', (vacation_id,)).fetchone()
    return jsonify(dict(row))


@app.route('/api/vacations/<int:vacation_id>', methods=['DELETE'])
def api_delete_vacation(vacation_id):
    db = get_db()
    db.execute('DELETE FROM vacations WHERE id = ?', (vacation_id,))
    db.commit()
    return jsonify({'status': 'deleted'})


@app.route('/api/calendar/<int:year>', methods=['GET'])
def api_get_calendar(year):
    config = get_config()
    events = generate_calendar_events(year, config)
    return jsonify({'year': year, 'events': events})


@app.route('/api/calendar/<int:year>/<int:month>', methods=['GET'])
def api_get_month_calendar(year, month):
    config = get_config()
    events = []
    us_holidays = get_us_holidays(year)
    for date, name in us_holidays.items():
        if date.month == month:
            events.append({
                'date': date.strftime('%Y-%m-%d'),
                'type': 'holiday',
                'name': name,
                'color': '#e74c3c'
            })
    db = get_db()
    last_day = calendar.monthrange(year, month)[1]
    rows = db.execute(
        'SELECT * FROM vacations WHERE start_date <= ? AND end_date >= ?',
        (f'{year}-{month:02d}-{last_day:02d}', f'{year}-{month:02d}-01')
    ).fetchall()
    for row in rows:
        start = datetime.strptime(row['start_date'], '%Y-%m-%d').date()
        end = datetime.strptime(row['end_date'], '%Y-%m-%d').date()
        month_start = datetime(year, month, 1).date()
        if month == 12:
            month_end = datetime(year, 12, 31).date()
        else:
            month_end = datetime(year, month + 1, 1).date() - timedelta(days=1)
        if start <= month_end and end >= month_start:
            overlap_start = max(start, month_start)
            overlap_end = min(end, month_end)
            overlap_start_text = overlap_start.strftime('%Y-%m-%d')
            events.append({
                # Keep date for clients that only inspect the legacy event shape.
                'date': overlap_start_text,
                'start_date': overlap_start_text,
                'end_date': overlap_end.strftime('%Y-%m-%d'),
                'type': 'vacation',
                'name': row['name'],
                'color': '#3498db',
                'vacation_id': row['id']
            })
    events.sort(key=lambda x: x['date'])
    return jsonify({'year': year, 'month': month, 'events': events})


@app.route('/api/stats', methods=['GET'])
def api_get_stats():
    config = get_config()
    today_date = datetime.now().date()
    today = today_date.strftime('%Y-%m-%d')
    balance = calculate_balance_on_date(today, config)
    forecast = generate_yearly_forecast(datetime.now().year, config)
    db = get_db()
    vacations = db.execute('SELECT * FROM vacations ORDER BY start_date').fetchall()
    year_end = date(today_date.year, 12, 31)
    remaining_scheduled_days = 0
    remaining_scheduled_hours = 0
    hours_per_day = config.get('pto_hours_per_day', 8.0) or 8.0
    for v in vacations:
        start = datetime.strptime(v['start_date'], '%Y-%m-%d').date()
        end = datetime.strptime(v['end_date'], '%Y-%m-%d').date()
        if start > year_end or end < today_date:
            continue
        scheduled_start = max(start, today_date)
        overlap_end = min(end, year_end)
        total_days = get_vacation_days(v['start_date'], v['end_date'], config)
        remaining_days = get_vacation_days(
            scheduled_start.strftime('%Y-%m-%d'),
            overlap_end.strftime('%Y-%m-%d'),
            config
        )
        if total_days > 0 and remaining_days > 0:
            remaining_scheduled_days += v['days'] * (remaining_days / total_days)
        elif total_days == 0 and start >= today_date:
            # Preserve manually entered PTO for date ranges without weekdays.
            remaining_scheduled_days += v['days']
        if start >= today_date:
            remaining_scheduled_hours += v['hours']
    if config.get('pto_accrual_type') == 'hours':
        remaining_total = (remaining_scheduled_days * hours_per_day) + remaining_scheduled_hours
    else:
        remaining_total = remaining_scheduled_days + (remaining_scheduled_hours / hours_per_day)
    remaining_scheduled_pto_days = round(remaining_total, 2)
    return jsonify({
        'today': today,
        'current_balance': balance,
        'yearly_forecast': forecast,
        'upcoming_vacations': len([
            v for v in vacations
            if datetime.strptime(v['end_date'], '%Y-%m-%d').date() >= today_date
        ]),
        # Keep the legacy field while exposing an explicitly named dashboard metric.
        'remaining_vacation_days': remaining_scheduled_pto_days,
        'remaining_scheduled_pto_days': remaining_scheduled_pto_days,
        'total_vacations': len(vacations)
    })


with app.app_context():
    init_db()

from csv import writer as csv_writer
from io import BytesIO, StringIO
from flask import send_file
from openpyxl import Workbook


def _validate_note_payload(data):
    data = data or {}
    note_date = data.get('date')
    text = str(data.get('text', '')).strip()
    if not isinstance(note_date, str) or not note_date or not text:
        return None, 'date and text are required'
    try:
        parsed_date = datetime.strptime(note_date, '%Y-%m-%d')
    except (TypeError, ValueError):
        return None, 'date must use yyyy-mm-dd format'
    if parsed_date.strftime('%Y-%m-%d') != note_date:
        return None, 'date must use yyyy-mm-dd format'
    return {'date': note_date, 'text': text}, None


@app.route('/api/notes', methods=['GET'])
def api_get_notes():
    db = get_db()
    note_date = request.args.get('date')
    if note_date:
        rows = db.execute(
            'SELECT * FROM notes WHERE date = ? ORDER BY date DESC, id DESC',
            (note_date,)
        ).fetchall()
    else:
        rows = db.execute('SELECT * FROM notes ORDER BY date DESC, id DESC').fetchall()
    return jsonify([dict(row) for row in rows])


@app.route('/api/notes', methods=['POST'])
def api_add_note():
    payload, error = _validate_note_payload(request.get_json())
    if error:
        return jsonify({'error': error}), 400
    db = get_db()
    cursor = db.execute(
        'INSERT INTO notes (date, text) VALUES (?, ?)',
        (payload['date'], payload['text'])
    )
    db.commit()
    row = db.execute('SELECT * FROM notes WHERE id = ?', (cursor.lastrowid,)).fetchone()
    return jsonify(dict(row)), 201


@app.route('/api/notes/<int:note_id>', methods=['PUT'])
def api_update_note(note_id):
    db = get_db()
    existing = db.execute('SELECT * FROM notes WHERE id = ?', (note_id,)).fetchone()
    if not existing:
        return jsonify({'error': 'Note not found'}), 404
    data = request.get_json() or {}
    payload, error = _validate_note_payload({
        'date': data.get('date', existing['date']),
        'text': data.get('text', existing['text'])
    })
    if error:
        return jsonify({'error': error}), 400
    db.execute(
        'UPDATE notes SET date = ?, text = ? WHERE id = ?',
        (payload['date'], payload['text'], note_id)
    )
    db.commit()
    row = db.execute('SELECT * FROM notes WHERE id = ?', (note_id,)).fetchone()
    return jsonify(dict(row))


@app.route('/api/notes/<int:note_id>', methods=['DELETE'])
def api_delete_note(note_id):
    db = get_db()
    db.execute('DELETE FROM notes WHERE id = ?', (note_id,))
    db.commit()
    return jsonify({'status': 'deleted'})


def _export_rows():
    db = get_db()
    config = get_config()
    year = datetime.now().year
    today = datetime.now().date().strftime('%Y-%m-%d')
    return config, calculate_balance_on_date(today, config), db.execute(
        'SELECT name, start_date, end_date, days, hours FROM vacations ORDER BY start_date'
    ).fetchall(), generate_yearly_forecast(year, config), year


def _safe_export_value(value):
    if isinstance(value, str) and value.startswith(('=', '+', '-', '@')):
        return "'" + value
    return value


@app.route('/api/export/excel', methods=['GET'])
def api_export_excel():
    config, balance, vacations, forecast, year = _export_rows()
    workbook = Workbook()
    summary = workbook.active
    summary.title = 'Balance Summary'
    summary.append([_safe_export_value(value) for value in ['Metric', 'Value']])
    summary.append([_safe_export_value(value) for value in ['Current Balance', balance['balance']]])
    summary.append([_safe_export_value(value) for value in ['Accrued YTD', balance['accrued']]])
    summary.append([_safe_export_value(value) for value in ['Used YTD', balance['used']]])
    summary.append([_safe_export_value(value) for value in ['Carryover from prior year', balance['carry']]])

    schedule = workbook.create_sheet('Vacation Schedule')
    schedule.append([_safe_export_value(value) for value in ['Name', 'Start', 'End', 'Days', 'Hours']])
    for vacation in vacations:
        schedule.append([_safe_export_value(value) for value in vacation])

    forecast_sheet = workbook.create_sheet('Monthly Forecast')
    forecast_sheet.append([_safe_export_value(value) for value in ['Month', 'Accrued', 'Used', 'Balance']])
    for month in forecast:
        forecast_sheet.append([_safe_export_value(value) for value in [
            month['month_name'], month['accrued'], month['used'], month['balance']
        ]])

    config_sheet = workbook.create_sheet('Configuration')
    config_sheet.append([_safe_export_value(value) for value in ['Setting', 'Value']])
    for key, value in sorted(config.items()):
        config_sheet.append([_safe_export_value(value) for value in [key, value]])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f'pto-export-{year}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/api/export/csv', methods=['GET'])
def api_export_csv():
    _, _, vacations, _, year = _export_rows()
    output = StringIO()
    csv = csv_writer(output)
    csv.writerow([_safe_export_value(value) for value in ['Name', 'Start', 'End', 'Days', 'Hours']])
    csv.writerows([
        [_safe_export_value(value) for value in vacation]
        for vacation in vacations
    ])
    response = send_file(
        BytesIO(output.getvalue().encode('utf-8')),
        as_attachment=True,
        download_name=f'pto-vacations-{year}.csv',
        mimetype='text/csv'
    )
    return response


if __name__ == '__main__':
    debug = os.getenv('FLASK_DEBUG', 'false').strip().lower() in {'1', 'true', 'yes', 'on'}
    app.run(debug=debug, host='0.0.0.0', port=5000)
